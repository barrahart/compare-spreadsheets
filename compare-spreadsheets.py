'''
--------------------------------------------------------------------------
    Name:         compare-spreadsheets.py

    Author:       Barra Hart

    Repo:         https://github.com/barrahart/compare-spreadsheets

    Description:  Compares ID columns in two Excel files and puts
                  all IDs not in both files in a third file.
--------------------------------------------------------------------------
'''

from openpyxl import * # library for working with Excel files

# fancy
def line():
    print('---------------------------------------------------------------------\n')


def compare(projectsA, projectsB, missingProjectsArray, missingProjects):

    # fancy
    print("\n\n")
    # compare ID column in file 1 against each ID in file 2
    for projectA in projectsA.iter_rows(values_only=True):
        flag=False

        for projectB in projectsB.iter_rows(values_only=True):
            if projectA[0]==projectB[0]:
                flag=True

        # notify user of progress
        print(f"> Done with search for ID {projectA[0]}\n")

        if flag==False: # ID not in second file
            missingProjectsArray.append(projectA[0])

    # update file for missing IDs
    for i in range(len(missingProjectsArray)):
        missingProjects[f'A{i+1}'] = missingProjectsArray[i]


# fancy
print('\n---------------------------------------------------------------------')
print("::::::::::::::::::::: Compare Spreadsheets 1.0 ::::::::::::::::::::::")
print('---------------------------------------------------------------------')
user_continue=input(":::::::::::: Press ENTER to continue, or CTRL-C to exit :::::::::::::\n")

line()
# import 1st file
user_work1 = input("\n> Enter full path for Spreadsheet 1: ")
workbook1 = load_workbook(f'{user_work1}')
sheet1=workbook1.worksheets[0]

# import 2nd file
user_work2 = input("\n> Enter full path for Spreadsheet 2: ")
workbook2 = load_workbook(f'{user_work2}')
sheet2=workbook2.worksheets[0]

# import 3rd file
workbook3 = load_workbook('./missing.xlsx')
sheet3 = workbook3.worksheets[0]

# temporary storage for missing IDs
missing=[]



compare(sheet1, sheet2, missing, sheet3)

'''TEST
line()
print("MISSING PROJECTS:\n\n")
for id in missing:
    print(id)
line()
'''

# save changes to 3rd file
workbook3.save('./missing.xlsx')

# notification
print("\n\n> Done. Check 'missing.xlsx' for your missing IDs!")
line()
